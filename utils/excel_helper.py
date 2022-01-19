#!/usr/bin/python
# -*- coding:utf-8 -*-

import xlrd
import xlwt
import openpyxl


class XlsHelper(object):
    """
    xls文件处理小工具
    """

    @staticmethod
    def get_data_list(excel_file_path):
        """
        批量获取数据
        :param excel_file_path:
        :return: data_list
        """
        # 打开文件
        data = xlrd.open_workbook(excel_file_path)
        # 获取第一个sheet
        table = data.sheets()[0]
        # 获取总行数
        row_count = table.nrows
        # 获取总列数
        col_count = table.ncols
        for row in range(1, row_count):
            print(table.cell(row, 1))
            print(table.cell(row, 3))


class XlsxHelper(object):
    """
    xlsx表处理小工具
    """

    @staticmethod
    def get_data_list(excel_file_path):
        """
        批量获取数据
        :param excel_file_path:
        :return: data_list
        """
        # 打开文件
        book = openpyxl.load_workbook(excel_file_path)
        # 获取当前活跃的sheet(默认第一个)
        sheet = book.active
        # 获取总行数
        row_count = sheet.max_row
        # 获取总列数
        col_count = sheet.max_column
        return [(sheet.cell(row, 1).value, sheet.cell(row, 3).value) for row in range(2, row_count + 1)]


if __name__ == '__main__':
    helper = XlsxHelper()
    data_list = helper.get_data_list("tw.xlsx")



